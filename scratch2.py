import openpyxl

wb = openpyxl.load_workbook('database/login.xlsx')
ws = wb.active

for row in range(2, ws.max_row + 1):
    chucdanh = ws.cell(row=row, column=5).value
    role = ws.cell(row=row, column=10).value
    
    if chucdanh == 'Trực Đập' and role == 'user':
        ws.cell(row=row, column=10, value='dap')

wb.save('database/login.xlsx')
print("Successfully updated roles for Trực Đập to 'dap'")
