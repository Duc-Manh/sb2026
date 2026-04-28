import openpyxl

wb = openpyxl.load_workbook('database/login.xlsx')
ws = wb.active

# Read the values
row2_vals = [cell.value for cell in ws[2]]
row3_vals = [cell.value for cell in ws[3]]
row4_vals = [cell.value for cell in ws[4]]

# Write the values in new order
for col_idx, val in enumerate(row3_vals, start=1):
    ws.cell(row=2, column=col_idx, value=val)

for col_idx, val in enumerate(row4_vals, start=1):
    ws.cell(row=3, column=col_idx, value=val)

for col_idx, val in enumerate(row2_vals, start=1):
    ws.cell(row=4, column=col_idx, value=val)

# Fix IDs to be sequential
ws.cell(row=2, column=1, value=1)
ws.cell(row=3, column=1, value=2)
ws.cell(row=4, column=1, value=3)

wb.save('database/login.xlsx')
print("Successfully rearranged the Excel rows.")
